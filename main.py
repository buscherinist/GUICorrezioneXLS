#per compilare
#pyinstaller --onefile --noconsole main.py
import openpyxl
from openpyxl import load_workbook
import os
import tkinter as tk
from tkinter import messagebox  # Importa messagebox per la finestra di conferma
from tkinter import filedialog, scrolledtext, Toplevel

#variabili globali
file_soluzione_excel = ""

def center_window():
    # Aggiorna le dimensioni della finestra in base ai contenuti
    root.update_idletasks()  # Aggiorna il layout e le dimensioni

    # Ottieni le dimensioni della finestra
    width = root.winfo_width()
    height = root.winfo_height()

    # Ottieni le dimensioni dello schermo
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()

    # Calcola la posizione x e y per centrare la finestra
    x = (screen_width // 2) - (width // 2)

    # Ottieni l'altezza della barra delle applicazioni (Windows)
    taskbar_height = 30

    y = (screen_height // 2) - (height // 2) - taskbar_height

    # Imposta la geometria della finestra
    root.geometry(f'{width}x{height}+{x}+{y}')

# Funzione di conferma per uscire dall'applicazione
def exit_app():
    if messagebox.askokcancel("Exit", "Vuoi uscire?"):
        root.quit()


# Funzione per cambiare il colore del pulsante quando il mouse entra
def on_enter(e):
    e.widget['background'] = colore_sfondo_button_on
    e.widget['foreground'] = colore_testo_button_on

# Funzione per ripristinare il colore del pulsante quando il mouse esce
def on_leave(e):
    e.widget['background'] = colore_sfondo_button
    e.widget['foreground'] = colore_testo_button


def mostra_file():
    file_path = "correzione.txt"  # Nome del file nella stessa cartella di main.py
    try:
        with open(file_path, "r") as file:
            contenuto = file.read()

        # Creazione di una finestra secondaria
        finestra_secondaria = Toplevel(root)
        finestra_secondaria.title(f"Contenuto di {file_path}")

        finestra_testo = scrolledtext.ScrolledText(finestra_secondaria, width=50, height=20)
        finestra_testo.pack(expand=True, fill="both")
        finestra_testo.insert("1.0", contenuto)

        # Impedisce l'editing del contenuto
        finestra_testo.config(state="disabled")

    except FileNotFoundError:
        print(f"Il file '{file_path}' non è stato trovato.")

def carica_soluzioni(file_soluzioni):
    # Legge il file di testo con le soluzioni e i punti, organizzati per foglio
    soluzioni = {}
    with open(file_soluzioni, 'r') as file:
        linee = file.readlines()

    foglio_corrente = None
    i = 0
    while i < len(linee):
        line = linee[i].strip()

        # Se la linea è un nome di foglio
        if line.startswith("esercizio"):
            foglio_corrente = line
            soluzioni[foglio_corrente] = {}
            i += 1
            continue

        # Altrimenti, leggiamo cella, formula e punti
        if foglio_corrente:
            cella = line
            formula = linee[i + 1].strip()
            punti = int(linee[i + 2].strip())
            soluzioni[foglio_corrente][cella] = (formula, punti)
            i += 3

    return soluzioni

def controlla_formule(nome_file_excel, soluzioni, file_output):
    # Apre il file Excel e controlla le formule
    workbook = openpyxl.load_workbook(nome_file_excel, data_only=False)
    punteggio_totale = 0
    risultati = {nome_file_excel: {}}

    file_output.write(f"File: {nome_file_excel}\n")
    for foglio_nome, celle in soluzioni.items():
        foglio = workbook[foglio_nome]
        file_output.write(f"  Foglio: {foglio_nome}\n")
        risultati[nome_file_excel][foglio_nome] = {}

        for cella, (formula_attesa, punti) in celle.items():
            valore_cella = foglio[cella].value

            if isinstance(valore_cella, str):
                valore_cella = valore_cella.replace(" ", "")

            if valore_cella == formula_attesa:
            #if uguali:
                risultati[nome_file_excel][foglio_nome][cella] = f"Formula corretta: {formula_attesa} (+{punti} punti)"
                file_output.write(f"    Cella {cella}: Formula corretta: {formula_attesa} (+{punti} punti)\n")
                punteggio_totale += punti
            else:
                risultati[nome_file_excel][foglio_nome][
                    cella] = f"Formula errata. Attesa: {formula_attesa}, Trovata: {valore_cella} (0 punti)"
                file_output.write(
                    f"    Cella {cella}: Formula errata. Attesa: {formula_attesa}, Trovata: {valore_cella} (0 punti)\n")

    file_output.write(f"\nPunteggio totale per {nome_file_excel}: {punteggio_totale}\n\n\n")
    return risultati, punteggio_totale

def calcola_punteggio_totale(file_elenco_excel, file_soluzioni, file_output):
    # Carica le soluzioni e inizializza il punteggio complessivo
    soluzioni = carica_soluzioni(file_soluzioni)
    risultati_globale = {}

    # Legge la lista dei file Excel
    with open(file_elenco_excel, 'r') as file:
        nomi_file_excel = [line.strip().lower()+".xlsx" for line in file if line.strip()]
    # Calcola il punteggio per ciascun file Excel
    for nome_file_excel in nomi_file_excel:
        risultati, punteggio_totale = controlla_formule(directory_verifiche+nome_file_excel, soluzioni, file_output)
        risultati_globale.update(risultati)

    return risultati_globale, punteggio_totale

def correggi():

    # Controlla se il file esiste
    if os.path.isfile(file_correzione):
        os.remove(file_correzione)  # Cancella il file

    with open(file_correzione, "a") as file_output:
        file_output.write("Correzione\n")
        calcola_punteggio_totale(file_elenco_excel, file_soluzioni, file_output)
    mostra_file()

def scegli_file():
    global file_soluzione_excel
    # Apri la finestra di dialogo per scegliere un file
    percorso_file = filedialog.askopenfilename()
    # Mostra il percorso del file selezionato (se esiste) in una Label
    if percorso_file:
        if percorso_file:
            nome_file = os.path.basename(percorso_file)
            label_path01.config(text=nome_file)
        file_soluzione_excel=percorso_file

def crea_soluzione():
    global file_soluzione_excel

    testo_label=label_path01.cget("text")

    if not testo_label.endswith(".xlsx"):
        messagebox.showinfo("Conferma", "Indicare prima il file delle soluzioni!")
        return

    # Percorso del file Excel
    percorso_file = file_soluzione_excel
    workbook = load_workbook(percorso_file)

    # Apre il file di output in modalità scrittura
    with open(file_soluzioni, "w") as f:
        # Itera su tutti i fogli del file
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Flag per scrivere il nome del foglio solo se ha celle colorate
            foglio_con_celle_colorate = False
            for row in sheet.iter_rows():
                for cell in row:
                    fill = cell.fill
                    # Controlla se la cella ha un colore diverso dal default
                    if fill and fill.start_color.index not in ["00000000", "FFFFFFFF","FBE2D5"]:
                        if not foglio_con_celle_colorate:
                            # Scrive il nome del foglio nel file
                            f.write(f"{sheet_name}\n")
                            foglio_con_celle_colorate = True

                        # Ottiene la formula della cella, se presente
                        formula = cell.value if cell.data_type == "f" else "Nessuna formula"

                        # Scrive nel file la coordinata e la formula della cella
                        f.write(f"{cell.coordinate}\n")
                        #f.write(f"{fill.start_color.index}\n")
                        f.write(f"{formula}\n")
                        f.write(f"1\n")

    #print("Risultati salvati in:", file_soluzioni)
    messagebox.showinfo("Conferma", "File delle soluzioni creato con successo!")

#main
#definisce i file di lavoro
file_elenco_excel = 'elencoalunni.txt'
file_soluzioni = 'soluzioni.txt'
file_correzione = "correzione.txt"
directory_verifiche = "./verifiche/"


#Definisce i colori
colore_sfondo_root = "#107C41"
colore_sfondo_line = "white"
colore_sfondo_titolo = "#5F70E7"
colore_testo_titolo = "white"
colore_sfondo_label = "white"
colore_testo_label = "black"
colore_sfondo_button = "white"
colore_testo_button = "black"
colore_sfondo_button_on = "#FF3E3E"
colore_testo_button_on = "white"
colore_sfondo_entry = "white"
colore_testo_entry = "black"
# Imposta il font
font_titolo = ("Georgia", 18)  # Font Arial, dimensione 16
font = ("Georgia", 16)  # Font Arial, dimensione 16
font_developed = ("Arial", 10)  # Font Arial, dimensione 16

# Crea la finestra principale
root = tk.Tk()
root.title("Correttore verifiche di Excel")
root.configure(bg=colore_sfondo_root)
larghezza_finestra=800
altezza_finestra=400
root.geometry(f"{larghezza_finestra}x{altezza_finestra}")
center_window()
# Sovrascrivi il comportamento del pulsante di chiusura
root.protocol("WM_DELETE_WINDOW", exit_app)

# Crea un frame
frame_center = tk.Frame(root, bg=colore_sfondo_root)
frame_center.pack(expand=True, fill="both")

# Etichetta vuota per lo spazio
spazio_vuoto = tk.Label(frame_center, text="", bg=colore_sfondo_root)
spazio_vuoto.grid(row=0, column=4, pady=10)  # Aggiungi una riga vuota con un padding verticale

# pulsante per aprire la finestra di dialogo
label_0 = tk.Label(frame_center, text="Scegli file delle soluzioni", bg=colore_sfondo_label, fg=colore_testo_label, font=font)
label_0.grid(row=1, column=0, padx=5, pady=2, sticky="w")

btn_scegli_file = tk.Button(frame_center, text="Scegli", comm=scegli_file, font=font)
btn_scegli_file.grid(row=1, column=1, pady=2, sticky="w")

# Label per mostrare il percorso del file selezionato
label_path01 = tk.Label(frame_center, text="Nessun file selezionato", wraplength=400, font=font)
label_path01.grid(row=1, column=2, pady=2, sticky="w")

# Prima riga: label e campo di input centrato
label_1 = tk.Label(frame_center, text="Creazione soluzione", bg=colore_sfondo_label, fg=colore_testo_label, font=font)
label_1.grid(row=2, column=0, padx=5, pady=2, sticky="w")

button_soluzione = tk.Button(frame_center, text="Crea soluzione", bg=colore_sfondo_button, fg=colore_testo_button, font=font, command=crea_soluzione)
button_soluzione.grid(row=2, column=1, pady=2, sticky="w")
# Bind degli eventi per il cambiamento di colore
button_soluzione.bind("<Enter>", on_enter)  # Quando il mouse entra nel pulsante
button_soluzione.bind("<Leave>", on_leave)  # Quando il mouse esce dal pulsante

# Prima riga: label e campo di input centrato
label_2 = tk.Label(frame_center, text="Correzione verifiche", bg=colore_sfondo_label, fg=colore_testo_label, font=font)
label_2.grid(row=3, column=0, padx=5, pady=2, sticky="w")

button_correggi = tk.Button(frame_center, text="Correggi", bg=colore_sfondo_button, fg=colore_testo_button, font=font, command=correggi)
button_correggi.grid(row=3, column=1, pady=2, sticky="w")
# Bind degli eventi per il cambiamento di colore
button_correggi.bind("<Enter>", on_enter)  # Quando il mouse entra nel pulsante
button_correggi.bind("<Leave>", on_leave)  # Quando il mouse esce dal pulsante

label_21 = tk.Label(frame_center, text="Ricorda: inserire le verifiche nella directory verifiche", wraplength=400, bg=colore_sfondo_label, fg=colore_testo_label, font=font)
label_21.grid(row=3, column=2, padx=5, pady=2, sticky="w")

# Etichetta vuota per lo spazio
spazio_vuoto = tk.Label(frame_center, text="", bg=colore_sfondo_root)
spazio_vuoto.grid(row=4, column=4, pady=10)  # Aggiungi una riga vuota con un padding verticale

#pulsante di chiusura
button_exit = tk.Button(frame_center, text="Exit",  bg=colore_sfondo_button, fg=colore_testo_button, font=font, command=exit_app)
button_exit.grid(row=5, column=2, pady=2)
# Bind degli eventi per il cambiamento di colore
button_exit.bind("<Enter>", on_enter)  # Quando il mouse entra nel pulsante
button_exit.bind("<Leave>", on_leave)  # Quando il mouse esce dal pulsante

# Avvia il loop principale della finestra
root.mainloop()